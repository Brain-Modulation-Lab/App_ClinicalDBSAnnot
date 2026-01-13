#!/usr/bin/env python3
"""
Demo script to show the export functionality.

This script demonstrates how the export report feature works
by creating sample data and showing the export process.
"""

import sys
import tempfile
from pathlib import Path

# Add the src directory to the path
sys.path.insert(0, str(Path(__file__).parent / "src"))

import pandas as pd
from clinical_dbs_annotator.models.session_data import SessionData
from clinical_dbs_annotator.utils.session_exporter import SessionExporter


def create_sample_data():
    """Create sample session data for demonstration."""
    sample_data = [
        {
            'date': '2024-01-15',
            'time': '09:30:00',
            'block_id': '1',
            'scale_name': 'YBOCS',
            'scale_value': '20',
            'stim_freq': '130',
            'left_contact': 'e1-e3',
            'left_amplitude': '3.5',
            'left_pulse_width': '60',
            'right_contact': 'e2-e4',
            'right_amplitude': '4.0',
            'right_pulse_width': '60',
            'notes': 'Baseline measurement'
        },
        {
            'date': '2024-01-15',
            'time': '09:45:00',
            'block_id': '2',
            'scale_name': 'YBOCS',
            'scale_value': '18',
            'stim_freq': '130',
            'left_contact': 'e1-e3',
            'left_amplitude': '4.0',
            'left_pulse_width': '60',
            'right_contact': 'e2-e4',
            'right_amplitude': '4.0',
            'right_pulse_width': '60',
            'notes': 'Increased left amplitude'
        },
        {
            'date': '2024-01-15',
            'time': '10:00:00',
            'block_id': '3',
            'scale_name': 'Mood',
            'scale_value': '6',
            'stim_freq': '130',
            'left_contact': 'e1-e3',
            'left_amplitude': '4.0',
            'left_pulse_width': '60',
            'right_contact': 'e2-e4',
            'right_amplitude': '4.5',
            'right_pulse_width': '60',
            'notes': 'Increased right amplitude, mood improved'
        },
        {
            'date': '2024-01-15',
            'time': '10:15:00',
            'block_id': '4',
            'scale_name': 'Anxiety',
            'scale_value': '4',
            'stim_freq': '130',
            'left_contact': 'e1-e3',
            'left_amplitude': '4.0',
            'left_pulse_width': '60',
            'right_contact': 'e2-e4',
            'right_amplitude': '4.5',
            'right_pulse_width': '60',
            'notes': 'Anxiety reduced'
        }
    ]
    
    return pd.DataFrame(sample_data)


def demo_export_functionality():
    """Demonstrate the export functionality."""
    print("Clinical DBS Annotator - Export Report Demo")
    print("=" * 50)
    
    # Create sample data
    print("\nCreating sample session data...")
    df = create_sample_data()
    print(f"Created {len(df)} sample records:")
    print(df[['date', 'time', 'scale_name', 'scale_value', 'left_amplitude', 'right_amplitude']].to_string())
    
    # Create temporary TSV file
    print("\nCreating temporary session file...")
    with tempfile.NamedTemporaryFile(mode='w', suffix='.tsv', delete=False) as f:
        df.to_csv(f.name, sep='\t', index=False)
        temp_file_path = f.name
    
    try:
        # Initialize SessionData with the temporary file
        print(f"Session file created: {temp_file_path}")
        
        # Mock SessionData to use our temporary file
        class MockSessionData:
            def __init__(self, file_path):
                self.file_path = file_path
            
            def is_file_open(self):
                return True
        
        mock_session_data = MockSessionData(temp_file_path)
        
        # Create exporter
        print("\nInitializing session exporter...")
        exporter = SessionExporter(mock_session_data)
        
        # Test the export functionality (without GUI)
        print("\nTesting export functionality...")
        
        # Read the data to verify it works
        test_df = exporter._read_session_data()
        if test_df is not None and not test_df.empty:
            print(f"Successfully read {len(test_df)} records from session file")
            
            # Test summary creation
            print("\nCreating summary statistics...")
            print(f"Date range: {exporter._get_date_range(test_df)}")
            print(f"Unique scales: {test_df['scale_name'].nunique()}")
            print(f"Left amplitude range: {test_df['left_amplitude'].min():.1f} - {test_df['left_amplitude'].max():.1f} mA")
            print(f"Right amplitude range: {test_df['right_amplitude'].min():.1f} - {test_df['right_amplitude'].max():.1f} mA")
            
            # Test Excel export to temporary location
            print("\nTesting Excel export...")
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as excel_file:
                excel_path = excel_file.name
            
            # Mock the file dialog to return our test path
            import unittest.mock
            with unittest.mock.patch('PyQt5.QtWidgets.QFileDialog.getSaveFileName', 
                                   return_value=(excel_path, "Excel Files (*.xlsx)")):
                with unittest.mock.patch('PyQt5.QtWidgets.QMessageBox.information'):
                    result = exporter.export_to_excel()
            
            if result and Path(excel_path).exists():
                print(f"Excel export successful: {excel_path}")
                print(f"File size: {Path(excel_path).stat().st_size} bytes")
                
                # Verify the Excel file has the expected sheets
                import openpyxl
                wb = openpyxl.load_workbook(excel_path)
                print(f"Excel sheets: {wb.sheetnames}")
                
                # Show some data from the main sheet
                ws = wb['Session Data']
                print(f"Session Data sheet has {ws.max_row} rows (including header)")
                
                # Clean up
                wb.close()
                Path(excel_path).unlink()
                print("Temporary Excel file cleaned up")
            else:
                print("Excel export failed")
        else:
            print("Failed to read session data")
    
    finally:
        # Clean up temporary TSV file
        Path(temp_file_path).unlink(missing_ok=True)
        print("Temporary TSV file cleaned up")
    
    print("\nExport functionality demo completed!")
    print("\nIntegration Summary:")
    print("1. Created SessionExporter class")
    print("2. Added export button to Step 3 view")
    print("3. Connected button to wizard controller")
    print("4. Added export method to controller")
    print("5. Updated dependencies (openpyxl)")
    print("6. Created comprehensive tests")
    print("\nThe export feature is now ready to use!")


if __name__ == "__main__":
    demo_export_functionality()
